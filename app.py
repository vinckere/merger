import unicodedata

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter


def normalize_colname(col):
    """Normalize accented characters (e.g. È → é)"""
    return unicodedata.normalize("NFKD", col).encode("ascii", "ignore").decode().lower()

def detect_store_column(df):
    """Detects the store name column using 'brest' or 'tours' in values"""
    for col in df.columns:
        if df[col].astype(str).str.contains("brest|tours", case=False).any():
            return col
    raise ValueError("Store column not found")

def extract_audio_ca(df):
    store_col = detect_store_column(df)
    df = df[[store_col, "CA Généré Audio"]].copy()
    df.columns = ["MAGASIN", "CA Audio"]
    df["CA Audio"] = df["CA Audio"].round().astype(int)
    df = df.iloc[:-1]
    return df

def extract_objectifs(df, reference_stores):
    store_col = detect_store_column(df)

    # Normalize columns for matching
    normalized_cols = {normalize_colname(c): c for c in df.columns}
    ca_col = normalized_cols["ca genere (factures - avoirs)"]
    objectif_col = normalized_cols["objectif (du mois)"]

    df = df[[store_col, objectif_col, ca_col]].copy()
    df.rename(columns={
        store_col: "MAGASIN",
        objectif_col: "OBJECTIF Mensuel",
        ca_col: "CA Mensuel Généré"
    }, inplace=True)

    df["OBJECTIF Mensuel"] = (df["OBJECTIF Mensuel"] * 1000).astype(int)
    df["CA Mensuel Généré"] = (df["CA Mensuel Généré"].str.replace(",", ".").astype(float) * 1000)

    df = df.iloc[:-1]  # remove total line
    current_stores = set(df["MAGASIN"].str.lower().str.strip())
    reference_set = set(reference_stores.str.lower().str.strip())
    if current_stores != reference_set:
        raise ValueError("Store mismatch between files")
    return df

def extract_optique_stats(df, reference_stores):
    store_col = detect_store_column(df)
    df = df[[store_col, "Nb Vente Opt", "Panier Moyen", "% Garantie", "% Pack Confort", "PM Pack Confort"]].copy()

    df.rename(columns={store_col: "MAGASIN"}, inplace=True)
    df["NB devis validés \n/ Panier moyen 450"] = df["Nb Vente Opt"].astype(float).round().astype(int).astype(
        str) + "/" + df["Panier Moyen"].astype(float).round().astype(int).astype(str)
    df["% SOP 45%"] = df["% Garantie"].astype(float).round().astype(int)
    df["% Pack Confort"] = df["% Pack Confort"].astype(float).round().astype(int)
    df["PM Pack Confort"] = df["PM Pack Confort"].astype(float).round().astype(int)
    df["% Pack Confort / PM Pack Confort"] = df["% Pack Confort"].astype(str) + " / " + df["PM Pack Confort"].astype( str)

    # Final selection
    df = df[["MAGASIN", "NB devis validés \n/ Panier moyen 450", "% SOP 45%", "% Pack Confort / PM Pack Confort"]]

    df = df.iloc[:-1]  # remove total row

    current = set(df["MAGASIN"].str.lower().str.strip())
    reference = set(reference_stores.str.lower().str.strip())
    if current != reference:
        raise ValueError("Store mismatch between files")

    return df


def save_to_excel(df):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    df = df.rename(columns={"CA Audio": "CA Audio généré "})
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value="STATISTIQUES OPTIQUES DU ...  AU ...")
    title_cell.font = Font(size=24, bold=True)
    title_cell.fill = PatternFill("solid", fgColor="BDD7EE")
    title_cell.alignment = Alignment(horizontal="center")

    for col_num, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_num, value=col_name)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(size=16)

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=16)

    # Apply center alignment and wrapping
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    #Adjust row width
    for i, column_cells in enumerate(ws.iter_cols(min_row=2), start=1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = length + 6

    # Adjust row height
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 28

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            # Check if it's the SOP column
            if cell.column_letter == "F" and isinstance(cell.value, int):
                if cell.value < 45:
                    cell.font = Font(size=16, color="FF0000")  # red
                else:
                    cell.font = Font(size=16, color="0070C0")  # blue

    wb.save(output)
    output.seek(0)
    return output

st.title("Caroptic Stats")

files = st.file_uploader(
    "1️⃣ Synthèse CA audio Année N  ||   2️⃣ Positionnement   ||   3️⃣ Synthèse stats optique Année N",
    type="csv",
    accept_multiple_files=True
)
if files and len(files) == 3:
    try:
        audio_df = pd.read_csv(files[0], encoding="latin1", skiprows=2, sep=";")
        objectifs_df = pd.read_csv(files[1], encoding="latin1", skiprows=2, sep=";", decimal=",")
        optique_df = pd.read_csv(files[2], encoding="latin1", skiprows=2, sep=";", decimal=",")

        #Debug logs
        st.write("Preview for store detection:", objectifs_df.head())
        st.write("Audio columns:", audio_df.columns.tolist())
        st.write("Objectifs columns:", objectifs_df.columns.tolist())
        st.write("Optique columns:", optique_df.columns.tolist())

        audio_data = extract_audio_ca(audio_df)
        objectifs_data = extract_objectifs(objectifs_df, audio_data["MAGASIN"])
        optique_data = extract_optique_stats(optique_df, audio_data["MAGASIN"])

        merged = audio_data.merge(objectifs_data, on="MAGASIN").merge(optique_data, on="MAGASIN")

        st.write(merged)

        excel_file = save_to_excel(merged)
        st.download_button("Download Excel", excel_file, "CHIFFRES DU XX AU XX.xlsx")
    except Exception as e:
        st.error(f"Error: {e}")
else:
    st.warning("Veuillez déposer les documents demandés dans le bon ordre")
